#!/usr/bin/env python3
"""Sync SEAMUN I 2027 conference data from the official allocation matrix xlsx."""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / 'Downloads' / 'SEAMUN I 2027 _ Allocation Matrix (2).xlsx'

SHEET_TO_ABBREV = {
    'ECOSOC': 'ECOSOC', 'F1': 'F1', 'Press Corps': 'Press Corps', 'UNICEF': 'UNICEF',
    'EU Parli': 'EP', 'UNESCO': 'UNESCO', 'UNHRC': 'UNHRC', 'UNODC': 'UNODC',
    'UNSC': 'UNSC', 'UN Women': 'UN Women', 'DISEC': 'DISEC',
    'FWC - Stranger Things': 'FWC', 'HSC': 'HSC', 'INTERPOL': 'Interpol', 'WHO': 'WHO',
}

MAIN_KEY = {
    'ECOSOC': 'ECOSOC', 'F1': 'F1', 'Press Corps': 'Press Corps', 'UNICEF': 'UNICEF',
    'EU Parli': 'EP', 'EU Parliament': 'EP', 'UNESCO': 'UNESCO', 'UNHRC': 'UNHRC',
    'UNODC': 'UNODC', 'UNSC': 'UNSC', 'UN Women': 'UN Women', 'DISEC': 'DISEC',
    'FWC - Stranger Things': 'FWC', 'FWC': 'FWC', 'HSC': 'HSC', 'INTERPOL': 'Interpol', 'WHO': 'WHO',
}

COMMITTEE_PREFIX = {
    'ECOSOC': 'ECOSOC (Economic and Societal Council)',
    'F1': 'F1 (Formula One Committee)',
    'Press Corps': 'Press Corps',
    'UNICEF': 'UNICEF (ESL)',
    'EP': 'EP (European Union)',
    'UNESCO': 'UNESCO (UNESCO)',
    'UNHRC': 'UNHRC (ESL)',
    'UNODC': 'UNODC (United Nations Office on Drugs and Crime)',
    'UNSC': 'UNSC (CRISIS)',
    'UN Women': 'UN Women (United Nations Entity for Gender Equality and the Empowerment of Women)',
    'DISEC': 'DISEC (Disarmament and International Security Committee)',
    'FWC': 'FWC (CRISIS) — STRANGER THINGS',
    'HSC': 'HSC (CRISIS)',
    'Interpol': 'Interpol (International Police and Criminal Investigation Organization)',
    'WHO': 'WHO (World Health Organization)',
}

GRADE = {
    'ECOSOC': ('7–12 (Year 8–13)', None),
    'F1': ('7–12 (Year 8–13)', None),
    'Press Corps': ('7–12 (Year 8–13)', None),
    'UNICEF': ('9–12 (Year 10–13)', None),
    'EP': ('7–12 (Year 8–13)', None),
    'UNESCO': ('7–12 (Year 8–13)', None),
    'UNHRC': ('9–12 (Year 10–13)', 'mature topics'),
    'UNODC': ('9–12 (Year 10–13)', 'mature topics'),
    'UNSC': ('7–12 (Year 8–13)', None),
    'UN Women': ('9–12 (Year 10–13)', 'mature topics'),
    'DISEC': ('7–12 (Year 8–13)', None),
    'FWC': ('9–12 (Year 10–13)', None),
    'HSC': ('7–12 (Year 8–13)', None),
    'Interpol': ('9–12 (Year 10–13)', None),
    'WHO': ('9–12 (Year 10–13)', 'mature topics'),
}

ORDER = ['ECOSOC', 'F1', 'Press Corps', 'UNICEF', 'EP', 'UNESCO', 'UNHRC', 'UNODC', 'UNSC', 'UN Women', 'DISEC', 'FWC', 'HSC', 'Interpol', 'WHO']


def parse_agendas(ws):
    agendas = []
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if cells[0] == 'Delegation':
            break
        if cells[0] in ('Head Chair', 'Co-Chair', 'Deputy Chair') or (cells[0] and 'Chair' in str(cells[0])):
            continue
        if cells[0] == 'Agenda':
            txt = cells[1]
        elif cells[0] is None and cells[1]:
            txt = cells[1]
        else:
            continue
        if not txt:
            continue
        t = str(txt).strip()
        if 'Placard Code' in t or t.startswith('Name Surname'):
            continue
        if ' S&D =' in t:
            t = t.split(' S&D =')[0].strip()
        agendas.append(t)
    return agendas


def parse_delegations(ws):
    delegations, in_deleg = [], False
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if cells[0] == 'Delegation':
            in_deleg = True
            continue
        if in_deleg and cells[0]:
            name = str(cells[0]).strip()
            if name.lower().startswith('total'):
                break
            delegations.append(name)
    return delegations


def js_str(s):
    return json.dumps(s, ensure_ascii=False)


def build_update(wb):
    main_sizes = {}
    for r in wb['MAIN'].iter_rows(min_row=2, max_row=16, values_only=True):
        if r[0] and r[0] != 'TOTAL':
            key = MAIN_KEY.get(r[0], r[0])
            main_sizes[key] = {'delegates': int(r[1]), 'chairs': int(r[2])}

    parsed = {}
    for sheet, abbrev in SHEET_TO_ABBREV.items():
        ws = wb[sheet]
        agendas = parse_agendas(ws)
        delegations = parse_delegations(ws)
        m = main_sizes.get(abbrev, {})
        delegates = m.get('delegates', len(delegations))
        chairs = m.get('chairs', 2)
        topics = ' | '.join(agendas)
        prefix = COMMITTEE_PREFIX[abbrev]
        if abbrev == 'FWC':
            line = prefix + ' — ' + topics
        elif topics:
            line = prefix + ' - ' + topics
            if abbrev == 'Press Corps' and not line.endswith('.'):
                line += '.'
        else:
            line = prefix
        parsed[abbrev] = {
            'agendas': agendas,
            'delegations': delegations,
            'delegates': delegates,
            'chairs': chairs,
            'committee_line': line,
        }

    committee_sizes = []
    for abbrev in ORDER:
        p = parsed[abbrev]
        gr, gn = GRADE[abbrev]
        entry = {
            'abbrev': abbrev,
            'chairLabel': 'Editor in Chief, Editor' if abbrev == 'Press Corps' else 'Chairs',
            'chairs': p['chairs'],
            'delegates': p['delegates'],
            'total': p['delegates'] + p['chairs'],
            'gradeRange': gr,
        }
        if gn:
            entry['gradeNote'] = gn
        if abbrev == 'Press Corps':
            entry['participantLabel'] = 'Journalists'
        committee_sizes.append(entry)

    committees = [parsed[a]['committee_line'] for a in ORDER]
    allocations = [
        f"{abbrev} ({parsed[abbrev]['delegates']} {'journalists' if abbrev == 'Press Corps' else 'delegates'}): {', '.join(parsed[abbrev]['delegations'])}"
        for abbrev in ORDER
    ]
    unique_topics = []
    for abbrev in ORDER:
        for ag in parsed[abbrev]['agendas']:
            short = ag.replace('The Question of ', '').strip()
            unique_topics.append(short)

    return {
        'committeeSizes': committee_sizes,
        'committees': committees,
        'allocations': allocations,
        'uniqueTopics': unique_topics,
        'total_delegates': sum(parsed[a]['delegates'] for a in ORDER),
        'total_chairs': sum(parsed[a]['chairs'] for a in ORDER),
    }


def js_committee_sizes(sizes):
    lines = ['    committeeSizes: [']
    for s in sizes:
        lines.append('        {')
        lines.append(f'            abbrev: {js_str(s["abbrev"])},')
        lines.append(f'            chairLabel: {js_str(s["chairLabel"])},')
        lines.append(f'            chairs: {s["chairs"]},')
        lines.append(f'            delegates: {s["delegates"]},')
        lines.append(f'            total: {s["total"]},')
        if s.get('gradeNote'):
            lines.append(f'            gradeRange: {js_str(s["gradeRange"])},')
            lines.append(f'            gradeNote: {js_str(s["gradeNote"])}')
        else:
            lines.append(f'            gradeRange: {js_str(s["gradeRange"])}')
        lines.append('        },')
    lines.append('    ],')
    return '\n'.join(lines)


def js_string_array(name, items, indent='    '):
    lines = [f'{indent}{name}: [']
    for item in items:
        lines.append(f'{indent}    {js_str(item)},')
    lines.append(f'{indent}],')
    return '\n'.join(lines)


def patch_conferences_data(js_path, update):
    text = js_path.read_text(encoding='utf-8')
    # Find id: 8 block - use markers
    start = text.find('    id: 8,')
    if start < 0:
        raise SystemExit('id: 8 not found')
    end = text.find('\n},\n{', start)
    if end < 0:
        end = text.find('\n}\n];', start)
    block = text[start:end]

    total_d = update['total_delegates']
    total_c = update['total_chairs']

    block = re.sub(
        r'description: "[^"]*"',
        f'description: "SEAMUN I 2027 — Policies with a Purpose. First student-led multinational independent MUN conference in the region (South East Asian Model United Nations Conference Thailand 2027). Two days in Bangkok with 15 committees across UN bodies, EU Parliament, Press Corps, F1, and crisis formats (UNSC, FWC, HSC). Non-profit conference; all profits donated to TRCS. {total_d} delegate positions, {total_c} chairs, 12 SMT, and 20+ advisors from across South East Asia."',
        block,
        count=1,
    )
    block = re.sub(
        r'size: "[^"]*"',
        f'size: "403+ ({total_d} delegates, {total_c} chairs, 12 SMT, 20+ advisors)"',
        block,
        count=1,
    )

    # Replace committeeSizes block
    block = re.sub(
        r'    committeeSizes: \[.*?\],',
        js_committee_sizes(update['committeeSizes']),
        block,
        flags=re.DOTALL,
    )

    # Replace committees array
    block = re.sub(
        r'    committees: \[.*?\],',
        js_string_array('committees', update['committees'], '    '),
        block,
        flags=re.DOTALL,
    )

    # Replace uniqueTopics
    block = re.sub(
        r'    uniqueTopics: \[.*?\],',
        js_string_array('uniqueTopics', update['uniqueTopics'], '    '),
        block,
        flags=re.DOTALL,
    )

    # Replace allocations
    block = re.sub(
        r'    allocations: \[.*?\],',
        js_string_array('allocations', update['allocations'], '    '),
        block,
        flags=re.DOTALL,
    )

    # Capacity line in extraNotes
    block = block.replace(
        '<strong>Capacity (target):</strong> 370 delegates, 33 chairs, 12 SMT, and 20+ advisors (435+ total).',
        f'<strong>Capacity (target):</strong> {total_d} delegates, {total_c} chairs, 12 SMT, and 20+ advisors (403+ total participants).',
    )

    new_text = text[:start] + block + text[end:]
    js_path.write_text(new_text, encoding='utf-8')


def patch_conferences_json(json_path, update):
    data = json.loads(json_path.read_text(encoding='utf-8'))
    for conf in data:
        if conf.get('id') == 8:
            conf['committeeSizes'] = update['committeeSizes']
            conf['committees'] = update['committees']
            conf['allocations'] = update['allocations']
            conf['uniqueTopics'] = update['uniqueTopics']
            td, tc = update['total_delegates'], update['total_chairs']
            conf['size'] = f'403+ ({td} delegates, {tc} chairs, 12 SMT, 20+ advisors)'
            conf['description'] = (
                f'SEAMUN I 2027 — Policies with a Purpose. First student-led multinational independent MUN conference in the region '
                f'(South East Asian Model United Nations Conference Thailand 2027). Two days in Bangkok with 15 committees across UN bodies, '
                f'EU Parliament, Press Corps, F1, and crisis formats (UNSC, FWC, HSC). Non-profit conference; all profits donated to TRCS. '
                f'{td} delegate positions, {tc} chairs, 12 SMT, and 20+ advisors from across South East Asia.'
            )
            extra = conf.get('extraNotes', '')
            extra = extra.replace(
                '<strong>Capacity (target):</strong> 370 delegates, 33 chairs, 12 SMT, and 20+ advisors (435+ total).',
                f'<strong>Capacity (target):</strong> {td} delegates, {tc} chairs, 12 SMT, and 20+ advisors (403+ total participants).',
            )
            conf['extraNotes'] = extra
            break
    json_path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + '\n', encoding='utf-8')


def main():
    if not XLSX.exists():
        print('xlsx not found:', XLSX, file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    update = build_update(wb)
    patch_conferences_data(ROOT / 'js' / 'conferences-data.js', update)
    patch_conferences_json(ROOT / 'data' / 'conferences.json', update)
    print(f"Updated SEAMUN I 2027: {update['total_delegates']} delegates, {update['total_chairs']} chairs, 15 committees")


if __name__ == '__main__':
    main()
